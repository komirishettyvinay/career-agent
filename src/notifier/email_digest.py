import smtplib
import json
import logging
import io
from datetime import date
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

import pandas as pd

from config.settings import GMAIL_SENDER, GMAIL_APP_PASSWORD, GMAIL_RECIPIENT
from src.storage.database import get_unemailed_jobs, mark_emailed

log = logging.getLogger(__name__)


def _build_excel(jobs: list[dict]) -> bytes:
    rows = []
    for j in jobs:
        matched  = json.loads(j.get("matched_keywords") or "[]")
        missing  = json.loads(j.get("missing_keywords") or "[]")
        skills   = json.loads(j.get("skills_required")  or "[]")
        rows.append({
            "Tier":              j.get("fit_tier", ""),
            "ATS Score":         j.get("ats_score", ""),
            "Role Name":         j.get("role_name", ""),
            "Title":             j.get("title", ""),
            "Company":           j.get("company", ""),
            "Location":          j.get("location", ""),
            "Salary":            j.get("salary", ""),
            "Skills Required":   ", ".join(skills),
            "Date Posted":       j.get("date_posted", ""),
            "Source":            j.get("source", ""),
            "Apply Link":        j.get("url", ""),
            "Matched Keywords":  ", ".join(matched),
            "Missing Keywords":  ", ".join(missing),
            "Summary":           j.get("ats_summary", ""),
        })

    df = pd.DataFrame(rows)

    # Sort: Strong first, then Maybe, then by score descending
    tier_order = {"Strong": 0, "Maybe": 1, "Skip": 2}
    df["_sort"] = df["Tier"].map(tier_order).fillna(3)
    df = df.sort_values(["_sort", "ATS Score"], ascending=[True, False]).drop(columns=["_sort"])

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Job Matches")

        ws = writer.sheets["Job Matches"]

        # Column widths
        col_widths = {
            "A": 10, "B": 12, "C": 28, "D": 28, "E": 25, "F": 20,
            "G": 25, "H": 14, "I": 35, "J": 14, "K": 50, "L": 35,
            "M": 35, "N": 50,
        }
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

        # Color rows by tier
        from openpyxl.styles import PatternFill, Font, Alignment
        green  = PatternFill("solid", fgColor="C6EFCE")
        yellow = PatternFill("solid", fgColor="FFEB9C")
        header_font = Font(bold=True)

        for cell in ws[1]:
            cell.font = header_font

        for row in ws.iter_rows(min_row=2):
            tier_val = row[0].value
            fill = green if tier_val == "Strong" else (yellow if tier_val == "Maybe" else None)
            for cell in row:
                if fill:
                    cell.fill = fill
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    return buf.getvalue()


def _build_html(jobs: list[dict]) -> str:
    strong = [j for j in jobs if j["fit_tier"] == "Strong"]
    maybe  = [j for j in jobs if j["fit_tier"] == "Maybe"]

    def job_block(job: dict) -> str:
        missing = json.loads(job.get("missing_keywords") or "[]")
        missing_str = ", ".join(missing[:5]) if missing else "None"
        return f"""
        <tr>
          <td style="padding:12px;border-bottom:1px solid #eee;">
            <b><a href="{job['url']}" style="color:#1a73e8;text-decoration:none;">{job['title']}</a></b><br>
            <span style="color:#555;">{job['company']} &bull; {job['location']}</span><br>
            <span style="font-size:13px;">ATS Score: <b>{job['ats_score']}/100</b> &bull; Missing: {missing_str}</span><br>
            <span style="font-size:12px;color:#888;">{job['ats_summary']}</span>
          </td>
        </tr>"""

    def section(title: str, jobs_list: list[dict]) -> str:
        if not jobs_list:
            return ""
        rows = "".join(job_block(j) for j in jobs_list)
        return f"""
        <h3 style="color:#333;margin-top:24px;">{title} ({len(jobs_list)})</h3>
        <table width="100%" cellpadding="0" cellspacing="0" style="border-collapse:collapse;">
          {rows}
        </table>"""

    today = date.today().strftime("%B %d, %Y")

    return f"""
    <html><body style="font-family:Arial,sans-serif;max-width:700px;margin:auto;color:#222;">
      <div style="background:#1a73e8;padding:20px;border-radius:8px 8px 0 0;">
        <h2 style="color:#fff;margin:0;">Job Hunter Daily Digest</h2>
        <p style="color:#c8deff;margin:4px 0 0;">{today} &bull; {len(jobs)} new matches &bull; Excel attached</p>
      </div>
      <div style="padding:20px;background:#fafafa;border:1px solid #eee;border-top:none;border-radius:0 0 8px 8px;">
        {section("🟢 Strong Matches (75+)", strong)}
        {section("🟡 Maybe (50-74)", maybe)}
        <hr style="margin:24px 0;border:none;border-top:1px solid #eee;">
        <p style="font-size:12px;color:#999;">
          Full details with apply links are in the attached Excel file.<br>
          Open your <a href="http://localhost:8501">Job Hunter Dashboard</a> for ATS gap analysis and Q&amp;A.
        </p>
      </div>
    </body></html>"""


def send_digest():
    if not GMAIL_APP_PASSWORD:
        log.warning("GMAIL_APP_PASSWORD not set — skipping email.")
        return

    jobs = get_unemailed_jobs()
    if not jobs:
        log.info("No new scored jobs to email.")
        return

    today = date.today().strftime("%b %d, %Y")
    subject = f"Job Hunter Digest — {today} ({len(jobs)} matches)"

    msg = MIMEMultipart("mixed")
    msg["Subject"] = subject
    msg["From"] = GMAIL_SENDER
    msg["To"] = GMAIL_RECIPIENT

    # HTML body
    msg.attach(MIMEText(_build_html(jobs), "html"))

    # Excel attachment
    excel_bytes = _build_excel(jobs)
    filename = f"jobs_{date.today().strftime('%Y-%m-%d')}.xlsx"
    part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    part.set_payload(excel_bytes)
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f'attachment; filename="{filename}"')
    msg.attach(part)

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(GMAIL_SENDER, GMAIL_APP_PASSWORD)
            server.sendmail(GMAIL_SENDER, GMAIL_RECIPIENT, msg.as_string())
        mark_emailed([j["job_hash"] for j in jobs])
        log.info(f"Email sent with Excel ({len(jobs)} jobs) to {GMAIL_RECIPIENT}")
    except Exception as e:
        log.error(f"Email send failed: {e}")
